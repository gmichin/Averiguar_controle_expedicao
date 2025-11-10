import pandas as pd
import numpy as np
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def limpar_valor_monetario(valor):
    """
    Limpa valores monetários removendo pontos e convertendo vírgulas para pontos
    """
    if pd.isna(valor) or valor == '':
        return 0.0
    
    valor_str = str(valor).strip()
    
    # Verifica se é negativo
    eh_negativo = valor_str.startswith('-')
    if eh_negativo:
        valor_str = valor_str[1:]  # Remove o sinal de negativo temporariamente
    
    valor_str = valor_str.replace('R$', '').replace(' ', '').strip()
    
    if valor_str.count(',') > 1:
        partes = valor_str.split(',')
        parte_inteira = ''.join(partes[:-1])
        parte_decimal = partes[-1]
        valor_limpo = parte_inteira + '.' + parte_decimal
    elif ',' in valor_str and '.' in valor_str:
        partes = valor_str.split(',')
        parte_inteira = partes[0].replace('.', '')
        valor_limpo = parte_inteira + '.' + partes[1]
    elif ',' in valor_str:
        partes = valor_str.split(',')
        if len(partes) == 2:
            parte_inteira = partes[0].replace('.', '')
            parte_decimal = partes[1]
            if len(parte_decimal) == 1:
                parte_decimal += '0'
            valor_limpo = parte_inteira + '.' + parte_decimal
        else:
            valor_limpo = valor_str.replace(',', '')
    else:
        valor_limpo = valor_str
    
    try:
        resultado = float(valor_limpo)
        if resultado > 1000000:
            valor_str_sem_virgulas = str(valor).replace(',', '').replace('.', '')
            if len(valor_str_sem_virgulas) > 2:
                resultado = float(valor_str_sem_virgulas[:-2] + '.' + valor_str_sem_virgulas[-2:])
        
        # Aplica o sinal negativo se necessário
        if eh_negativo:
            resultado = -resultado
            
        return resultado
    except:
        return 0.0

def limpar_valor_numerico(valor):
    """
    Limpa valores numéricos (peso, quantidade, etc.) preservando o sinal negativo
    """
    if pd.isna(valor) or valor == '':
        return 0.0
    
    valor_str = str(valor).strip()
    
    # Verifica se é negativo
    eh_negativo = valor_str.startswith('-')
    if eh_negativo:
        valor_str = valor_str[1:]  # Remove o sinal de negativo temporariamente
    
    # Mantém apenas dígitos, vírgulas, pontos e preserva o sinal negativo
    valor_str = ''.join(c for c in valor_str if c.isdigit() or c in ',.')
    
    if valor_str.count(',') > 1:
        partes = valor_str.split(',')
        parte_inteira = ''.join(partes[:-1])
        parte_decimal = partes[-1]
        valor_limpo = parte_inteira + '.' + parte_decimal
    elif ',' in valor_str and '.' in valor_str:
        partes = valor_str.split(',')
        parte_inteira = partes[0].replace('.', '')
        valor_limpo = parte_inteira + '.' + partes[1]
    elif ',' in valor_str:
        partes = valor_str.split(',')
        if len(partes) == 2:
            parte_inteira = partes[0].replace('.', '')
            parte_decimal = partes[1]
            if len(parte_decimal) == 1:
                parte_decimal += '0'
            valor_limpo = parte_inteira + '.' + parte_decimal
        else:
            valor_limpo = valor_str.replace(',', '')
    else:
        valor_limpo = valor_str
    
    try:
        resultado = float(valor_limpo)
        if resultado > 100000:
            valor_str_sem_virgulas = str(valor).replace(',', '').replace('.', '')
            if len(valor_str_sem_virgulas) > 2:
                resultado = float(valor_str_sem_virgulas[:-2] + '.' + valor_str_sem_virgulas[-2:])
        
        # Aplica o sinal negativo se necessário
        if eh_negativo:
            resultado = -resultado
            
        return resultado
    except:
        return 0.0

def comparar_valores_com_tolerancia(valor1, valor2, tolerancia=0.014):
    """
    Compara dois valores com tolerância de 0,014 para diferenças de ponto flutuante
    """
    try:
        valor1_float = float(valor1)
        valor2_float = float(valor2)
        return abs(valor1_float - valor2_float) <= tolerancia
    except:
        return False

def detectar_separador_csv(caminho):
    """
    Detecta o separador do arquivo CSV
    """
    try:
        with open(caminho, 'r', encoding='utf-8') as f:
            primeira_linha = f.readline()
        if ';' in primeira_linha:
            return ';'
        elif ',' in primeira_linha:
            return ','
        else:
            return ';'
    except:
        return ';'

def formatar_nota_fiscal(valor):
    """
    Formata a nota fiscal corretamente
    """
    if pd.isna(valor) or valor == '':
        return ''
    
    valor_str = str(valor).strip()
    valor_limpo = ''.join(c for c in valor_str if c.isdigit())
    
    if len(valor_limpo) == 7 and valor_limpo.endswith('0'):
        valor_limpo = valor_limpo[:-1]
    
    return valor_limpo

def converter_para_inteiro_nota_fiscal(valor):
    """
    Converte a nota fiscal para inteiro
    """
    if pd.isna(valor) or valor == '':
        return 0
    
    valor_formatado = formatar_nota_fiscal(valor)
    try:
        return int(valor_formatado) if valor_formatado else 0
    except:
        return 0

def ler_csv_com_cabecalho(caminho, data_inicio=None, data_fim=None):
    """
    Lê o CSV detectando automaticamente o cabeçalho e filtrando por data
    E APENAS NOTAS COM HISTÓRICO 51
    """
    separador = detectar_separador_csv(caminho)
    
    try:
        df_teste = pd.read_csv(caminho, nrows=0, sep=separador, encoding='utf-8')
        colunas_disponiveis = df_teste.columns.tolist()
        
        mapeamento_colunas = {
            'NOTA FISCAL': ['NF-E', 'NOTA FISCAL', 'NOTAFISCAL', 'NOTA_FISCAL', 'NF', 'NOTA'],
            'PESO': ['QTDE REAL', 'QTDE_REAL', 'QUANTIDADE REAL', 'PESO', 'PESO_KG', 'PESO KG', 'PESO_TOTAL', 'QUANTIDADE', 'QTDE'],
            'TOTAL': ['FAT BRUTO', 'FAT_BRUTO', 'TOTAL', 'TOTAL_NF', 'VALOR_TOTAL', 'VALOR TOTAL', 'VALOR'],
            'DATA': ['DATA', 'DATE', 'DT', 'DATA_NF', 'DATA EMISSÃO', 'EMISSÃO'],
            'HISTÓRICO': ['HISTÓRICO', 'HISTORICO', 'HIST', 'HISTORICO_LANCTO']
        }
        
        colunas_para_ler = []
        for coluna_base, alternativas in mapeamento_colunas.items():
            encontrou = False
            for alternativa in alternativas:
                for coluna_csv in colunas_disponiveis:
                    if coluna_csv.upper() == alternativa.upper():
                        colunas_para_ler.append(coluna_csv)
                        encontrou = True
                        print(f"   ✅ Coluna '{coluna_base}' encontrada como: '{coluna_csv}'")
                        break
                if encontrou:
                    break
        
        # Lê todo o CSV
        df = pd.read_csv(caminho, sep=separador, usecols=colunas_para_ler, encoding='utf-8')
        
        rename_dict = {}
        for coluna_csv in colunas_para_ler:
            for coluna_base, alternativas in mapeamento_colunas.items():
                if any(coluna_csv.upper() == alt.upper() for alt in alternativas):
                    rename_dict[coluna_csv] = coluna_base
                    break
        
        df = df.rename(columns=rename_dict)
        
        # FILTRO CRÍTICO: APENAS HISTÓRICO 51
        if 'HISTÓRICO' in df.columns:
            # Converte histórico para numérico e filtra apenas 51
            df['HISTÓRICO'] = pd.to_numeric(df['HISTÓRICO'], errors='coerce')
            df = df[df['HISTÓRICO'] == 51]
            print(f"   ✅ CSV filtrado - apenas histórico 51: {len(df)} notas")
        
        # FILTRO ADICIONAL: APENAS LINHAS COM QTDE REAL POSITIVA
        if 'PESO' in df.columns:
            df['PESO_LIMPO'] = df['PESO'].apply(limpar_valor_numerico)
            print(f"   📊 Valores únicos de QTDE REAL encontrados: {df['PESO_LIMPO'].unique()}")
            # Remove TODAS as linhas com valores negativos ou zero
            df = df[df['PESO_LIMPO'] > 0]
            print(f"   ✅ CSV filtrado - apenas QTDE REAL positiva: {len(df)} notas")
        
        if (data_inicio or data_fim) and 'DATA' in df.columns:
            df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce')
            
            if data_inicio:
                df = df[df['DATA'] >= data_inicio]
            if data_fim:
                data_fim_ajustada = data_fim + pd.Timedelta(days=1)
                df = df[df['DATA'] < data_fim_ajustada]
        
        return df
        
    except Exception as e:
        try:
            df_teste = pd.read_csv(caminho, nrows=0, sep=separador, encoding='latin-1')
            colunas_disponiveis = df_teste.columns.tolist()
            
            colunas_para_ler = []
            mapeamento_colunas = {
                'NOTA FISCAL': ['NF-E', 'NOTA FISCAL', 'NOTAFISCAL', 'NOTA_FISCAL', 'NF', 'NOTA'],
                'PESO': ['QTDE REAL', 'QTDE_REAL', 'QUANTIDADE REAL', 'PESO', 'PESO_KG', 'PESO KG', 'PESO_TOTAL', 'QUANTIDADE', 'QTDE'],
                'TOTAL': ['FAT BRUTO', 'FAT_BRUTO', 'TOTAL', 'TOTAL_NF', 'VALOR_TOTAL', 'VALOR TOTAL', 'VALOR'],
                'DATA': ['DATA', 'DATE', 'DT', 'DATA_NF', 'DATA EMISSÃO', 'EMISSÃO'],
                'HISTÓRICO': ['HISTÓRICO', 'HISTORICO', 'HIST', 'HISTORICO_LANCTO']
            }
            
            for coluna_base, alternativas in mapeamento_colunas.items():
                encontrou = False
                for alternativa in alternativas:
                    for coluna_csv in colunas_disponiveis:
                        if coluna_csv.upper() == alternativa.upper():
                            colunas_para_ler.append(coluna_csv)
                            encontrou = True
                            print(f"   ✅ Coluna '{coluna_base}' encontrada como: '{coluna_csv}'")
                            break
                    if encontrou:
                        break
            
            # Lê todo o CSV
            df = pd.read_csv(caminho, sep=separador, usecols=colunas_para_ler, encoding='latin-1')
            
            rename_dict = {}
            for coluna_csv in colunas_para_ler:
                for coluna_base, alternativas in mapeamento_colunas.items():
                    if any(coluna_csv.upper() == alt.upper() for alt in alternativas):
                        rename_dict[coluna_csv] = coluna_base
                        break
            
            df = df.rename(columns=rename_dict)
            
            # FILTRO CRÍTICO: APENAS HISTÓRICO 51
            if 'HISTÓRICO' in df.columns:
                # Converte histórico para numérico e filtra apenas 51
                df['HISTÓRICO'] = pd.to_numeric(df['HISTÓRICO'], errors='coerce')
                df = df[df['HISTÓRICO'] == 51]
                print(f"   ✅ CSV filtrado - apenas histórico 51: {len(df)} notas")
            
            # FILTRO ADICIONAL: APENAS LINHAS COM QTDE REAL POSITIVA
            if 'PESO' in df.columns:
                df['PESO_LIMPO'] = df['PESO'].apply(limpar_valor_numerico)
                print(f"   📊 Valores únicos de QTDE REAL encontrados: {df['PESO_LIMPO'].unique()}")
                # Remove TODAS as linhas com valores negativos ou zero
                df = df[df['PESO_LIMPO'] > 0]
                print(f"   ✅ CSV filtrado - apenas QTDE REAL positiva: {len(df)} notas")
            
            if (data_inicio or data_fim) and 'DATA' in df.columns:
                df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce')
                
                if data_inicio:
                    df = df[df['DATA'] >= data_inicio]
                if data_fim:
                    data_fim_ajustada = data_fim + pd.Timedelta(days=1)
                    df = df[df['DATA'] < data_fim_ajustada]
            
            return df
            
        except Exception as e2:
            print(f"❌ Erro ao ler CSV: {e2}")
            return None

def obter_periodo_usuario():
    """
    Solicita o período desejado ao usuário
    """
    print("📅 PERÍODO DE ANÁLISE")
    print("Deixe em branco para não usar filtro")
    
    while True:
        data_inicio_str = input("Data de início (DD/MM/AAAA): ").strip()
        data_fim_str = input("Data de fim (DD/MM/AAAA): ").strip()
        
        data_inicio = None
        data_fim = None
        
        try:
            if data_inicio_str:
                data_inicio = datetime.strptime(data_inicio_str, '%d/%m/%Y')
            if data_fim_str:
                data_fim = datetime.strptime(data_fim_str, '%d/%m/%Y')
            
            if data_inicio and data_fim and data_inicio > data_fim:
                print("❌ Data de início não pode ser maior que data de fim!")
                continue
                
            return data_inicio, data_fim
            
        except ValueError:
            print("❌ Formato inválido! Use DD/MM/AAAA")
            continuar = input("Tentar novamente? (S/N): ").strip().upper()
            if continuar != 'S':
                return None, None

def aplicar_estilo_erros(worksheet, df_relatorio, problemas_por_linha):
    """
    Aplica estilo vermelho às células com problemas graves e amarelo para erros de digitação
    """
    fill_vermelho = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
    fill_amarelo = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    
    colunas_indices = {}
    for idx, col_name in enumerate(df_relatorio.columns):
        colunas_indices[col_name] = idx
    
    for linha_idx, problemas in problemas_por_linha.items():
        linha_excel = linha_idx + 2
        
        for problema in problemas:
            tipo = problema['tipo']
            
            if tipo == 'nf_nao_encontrada':
                for col_idx in range(len(df_relatorio.columns)):
                    cell = worksheet.cell(row=linha_excel, column=col_idx+1)
                    cell.fill = fill_vermelho
            
            elif tipo == 'peso_divergente':
                if 'VOG_Expedição' in colunas_indices:
                    col_idx = colunas_indices['VOG_Expedição']
                    cell = worksheet.cell(row=linha_excel, column=col_idx+1)
                    cell.fill = fill_vermelho
                if 'PESO_CSV' in colunas_indices:
                    col_idx = colunas_indices['PESO_CSV']
                    cell = worksheet.cell(row=linha_excel, column=col_idx+1)
                    cell.fill = fill_vermelho
            
            elif tipo == 'valor_divergente':
                if 'R$ NF_Expedição' in colunas_indices:
                    col_idx = colunas_indices['R$ NF_Expedição']
                    cell = worksheet.cell(row=linha_excel, column=col_idx+1)
                    cell.fill = fill_vermelho
                if 'TOTAL_CSV' in colunas_indices:
                    col_idx = colunas_indices['TOTAL_CSV']
                    cell = worksheet.cell(row=linha_excel, column=col_idx+1)
                    cell.fill = fill_vermelho
            
            elif tipo == 'erro_digitacao_vog':
                if 'VOG_Expedição' in colunas_indices:
                    col_idx = colunas_indices['VOG_Expedição']
                    cell = worksheet.cell(row=linha_excel, column=col_idx+1)
                    cell.fill = fill_amarelo
            
            elif tipo == 'erro_digitacao_valor':
                if 'R$ NF_Expedição' in colunas_indices:
                    col_idx = colunas_indices['R$ NF_Expedição']
                    cell = worksheet.cell(row=linha_excel, column=col_idx+1)
                    cell.fill = fill_amarelo

def processar_planilhas():
    # Caminhos das planilhas
    caminho_expedicao = r"Z:\RODRIGO - LOGISTICA\Cópia de CONTROLE DE EXPEDIÇÃO NOVEMBRO.xlsx"
    caminho_csv = r"S:\hor\excel\fechamento-20251101-20251110.csv"
    
    if not os.path.exists(caminho_expedicao):
        print(f"❌ Arquivo não encontrado: {caminho_expedicao}")
        return
    
    if not os.path.exists(caminho_csv):
        print(f"❌ Arquivo não encontrado: {caminho_csv}")
        return
    
    # Solicita o período ao usuário
    data_inicio, data_fim = obter_periodo_usuario()
    
    print("\n📊 PROCESSANDO DADOS...")
    
    try:
        # Lê a planilha de controle de expedição
        df_expedicao = pd.read_excel(
            caminho_expedicao, 
            sheet_name='JAN-FEV-MAR-ABR-MAI-JUN',
            header=3,
            usecols=['NF', 'VOG', 'R$ NF', 'STATUS', 'DATA', 'OPERAÇÃO'],
            dtype={'NF': str}
        )
        
        df_expedicao = df_expedicao.dropna(subset=['NF'])
        
        status_validos = ['ENTREGUE', 'EM ROTA', 'DEVOLUÇÃO']
        df_expedicao_filtrado = df_expedicao[df_expedicao['STATUS'].isin(status_validos)].copy()
        
        operacoes_vog = ['VOG', 'VOG 2ºSAIDA', 'VOG 2 SAIDA', 'VOG 2SAIDA', 'VOG 2º SAIDA']
        df_expedicao_filtrado = df_expedicao_filtrado[
            df_expedicao_filtrado['OPERAÇÃO'].isin(operacoes_vog)
        ]
        
        df_expedicao_filtrado['NF'] = df_expedicao_filtrado['NF'].apply(formatar_nota_fiscal)
        
        if 'DATA' in df_expedicao_filtrado.columns:
            df_expedicao_filtrado['DATA_EXPEDICAO'] = pd.to_datetime(df_expedicao_filtrado['DATA'], errors='coerce')
            
            if data_inicio or data_fim:
                if data_inicio:
                    df_expedicao_filtrado = df_expedicao_filtrado[df_expedicao_filtrado['DATA_EXPEDICAO'] >= data_inicio]
                if data_fim:
                    data_fim_ajustada = data_fim + pd.Timedelta(days=1)
                    df_expedicao_filtrado = df_expedicao_filtrado[df_expedicao_filtrado['DATA_EXPEDICAO'] < data_fim_ajustada]
        else:
            df_expedicao_filtrado['DATA_EXPEDICAO'] = None
        
        vog_limpo = df_expedicao_filtrado['VOG'].apply(limpar_valor_numerico)
        valor_nf_limpo = df_expedicao_filtrado['R$ NF'].apply(limpar_valor_monetario)
        
        print(f"   ✅ Expedição processada: {len(df_expedicao_filtrado)} notas VOG")
        
    except Exception as e:
        print(f"❌ Erro ao ler planilha de expedição: {e}")
        return
    
    # Lê o arquivo CSV (AGORA APENAS COM HISTÓRICO 51 E QTDE REAL POSITIVA)
    print("   📋 Lendo arquivo CSV...")
    df_csv = ler_csv_com_cabecalho(caminho_csv, data_inicio, data_fim)
    
    if df_csv is None or df_csv.empty:
        print("❌ Não foi possível ler o arquivo CSV ou nenhum dado com histórico 51 e QTDE REAL positiva encontrado")
        return
    
    df_csv = df_csv.dropna(subset=['NOTA FISCAL'])
    df_csv['NOTA FISCAL'] = df_csv['NOTA FISCAL'].apply(formatar_nota_fiscal)
    
    # Usa o PESO_LIMPO que já foi calculado na função ler_csv_com_cabecalho
    peso_limpo = df_csv['PESO_LIMPO'] if 'PESO_LIMPO' in df_csv.columns else df_csv['PESO'].apply(limpar_valor_numerico)
    total_limpo = df_csv['TOTAL'].apply(limpar_valor_monetario)
    
    df_csv['PESO_COMPARACAO'] = peso_limpo
    df_csv['TOTAL_COMPARACAO'] = total_limpo
    
    if 'DATA' in df_csv.columns:
        df_csv['DATA_CSV'] = pd.to_datetime(df_csv['DATA'], dayfirst=True, errors='coerce')
    else:
        df_csv['DATA_CSV'] = None
    
    # AGRUPAMENTO POR NOTA FISCAL - SOMANDO APENAS VALORES POSITIVOS (já filtrados)
    df_agrupado = df_csv.groupby('NOTA FISCAL').agg({
        'PESO': 'first',
        'TOTAL': 'first',
        'DATA_CSV': 'first',
        'PESO_COMPARACAO': 'sum',  # Soma apenas os valores positivos (já filtrados)
        'TOTAL_COMPARACAO': 'sum'  # Soma apenas os valores positivos (já filtrados)
    }).reset_index()
    
    print(f"   ✅ CSV agrupado: {len(df_agrupado)} notas únicas (histórico 51 + QTDE REAL positiva - valores somados)")
    
    # DEBUG: Mostrar algumas notas para verificar se os valores estão corretos
    print("\n   🔍 VERIFICAÇÃO DE VALORES (amostra):")
    for i, row in df_agrupado.head(5).iterrows():
        print(f"      NF {row['NOTA FISCAL']}: PESO_COMPARACAO = {row['PESO_COMPARACAO']}, TOTAL_COMPARACAO = {row['TOTAL_COMPARACAO']}")
    
    # Realiza o merge das planilhas
    df_comparacao = pd.merge(
        df_expedicao_filtrado,
        df_agrupado,
        left_on='NF',
        right_on='NOTA FISCAL',
        how='left',
        indicator=True
    )
    
    df_csv_sem_expedicao = pd.merge(
        df_agrupado,
        df_expedicao_filtrado,
        left_on='NOTA FISCAL',
        right_on='NF',
        how='left',
        indicator=True
    )
    nfs_csv_sem_expedicao = df_csv_sem_expedicao[df_csv_sem_expedicao['_merge'] == 'left_only']
    
    # Identifica divergências
    resultados = []
    problemas_por_linha = {}
    
    for index, row in df_comparacao.iterrows():
        nf = row['NF']
        vog_expedicao = vog_limpo.iloc[index] if index < len(vog_limpo) else 0
        valor_nf_expedicao = valor_nf_limpo.iloc[index] if index < len(valor_nf_limpo) else 0
        peso_csv = row['PESO_COMPARACAO'] if 'PESO_COMPARACAO' in row and not pd.isna(row['PESO_COMPARACAO']) else 0
        total_csv = row['TOTAL_COMPARACAO'] if 'TOTAL_COMPARACAO' in row and not pd.isna(row['TOTAL_COMPARACAO']) else 0
        status = row['STATUS']
        operacao = row['OPERAÇÃO'] if 'OPERAÇÃO' in row else ''
        merge_status = row['_merge']
        
        data_expedicao = None
        data_csv = None
        
        if 'DATA_EXPEDICAO' in row and not pd.isna(row['DATA_EXPEDICAO']):
            data_expedicao = row['DATA_EXPEDICAO']
        
        if 'DATA_CSV' in row and not pd.isna(row['DATA_CSV']):
            data_csv = row['DATA_CSV']
        
        problemas_linha = []
        
        if merge_status == 'left_only':
            problemas_linha.append({'tipo': 'nf_nao_encontrada', 'descricao': 'NF não encontrada no CSV (histórico 51 + QTDE REAL positiva)'})
        else:
            # AGORA COM TOLERÂNCIA DE 0,014
            if not comparar_valores_com_tolerancia(vog_expedicao, peso_csv, 0.014):
                problemas_linha.append({'tipo': 'peso_divergente', 'descricao': f"PESO divergente: Expedição={vog_expedicao} vs CSV={peso_csv}"})
            
            # AGORA COM TOLERÂNCIA DE 0,014
            if not comparar_valores_com_tolerancia(valor_nf_expedicao, total_csv, 0.014):
                problemas_linha.append({'tipo': 'valor_divergente', 'descricao': f"VALOR divergente: Expedição={valor_nf_expedicao} vs CSV={total_csv}"})
        
        if pd.notna(row['VOG']):
            vog_original = str(row['VOG'])
            if vog_original.count(',') > 1 or (',' in vog_original and '.' in vog_original.split(',')[0]):
                problemas_linha.append({'tipo': 'erro_digitacao_vog', 'descricao': f"Erro digitação VOG"})
        
        if pd.notna(row['R$ NF']):
            valor_original = str(row['R$ NF'])
            if valor_original.count(',') > 1 or (',' in valor_original and '.' in valor_original.split(',')[0]):
                problemas_linha.append({'tipo': 'erro_digitacao_valor', 'descricao': f"Erro digitação VALOR"})
        
        if problemas_linha:
            nf_inteiro = converter_para_inteiro_nota_fiscal(nf)
            peso_csv_float = float(peso_csv) if peso_csv else 0.0
            total_csv_float = float(total_csv) if total_csv else 0.0
            
            resultados.append({
                'NF': nf_inteiro,
                'DATA_EXPEDICAO': data_expedicao,
                'DATA_CSV': data_csv,
                'STATUS': status,
                'OPERAÇÃO': operacao,
                'VOG_Expedição': row['VOG'],
                'PESO_CSV': peso_csv_float,
                'R$ NF_Expedição': row['R$ NF'],
                'TOTAL_CSV': total_csv_float
            })
            
            problemas_por_linha[len(resultados)-1] = problemas_linha
    
    for index, row in nfs_csv_sem_expedicao.iterrows():
        nf = row['NOTA FISCAL']
        data_csv = row['DATA_CSV'] if 'DATA_CSV' in row and not pd.isna(row['DATA_CSV']) else None
        
        peso_csv = row['PESO'] if 'PESO' in row else ''
        total_csv = row['TOTAL'] if 'TOTAL' in row else ''
        
        nf_inteiro = converter_para_inteiro_nota_fiscal(nf)
        peso_csv_float = limpar_valor_numerico(peso_csv) if peso_csv else 0.0
        total_csv_float = limpar_valor_monetario(total_csv) if total_csv else 0.0
        
        resultados.append({
            'NF': nf_inteiro,
            'DATA_EXPEDICAO': None,
            'DATA_CSV': data_csv,
            'STATUS': 'N/A',
            'OPERAÇÃO': 'N/A',
            'VOG_Expedição': 'N/A',
            'PESO_CSV': peso_csv_float,
            'R$ NF_Expedição': 'N/A',
            'TOTAL_CSV': total_csv_float
        })
        
        problemas_por_linha[len(resultados)-1] = [{'tipo': 'nf_nao_encontrada', 'descricao': 'NF do CSV (histórico 51 + QTDE REAL positiva) não encontrada no expedição'}]
    
    # Cria relatório final
    if resultados:
        df_relatorio = pd.DataFrame(resultados)
        
        colunas_ordenadas = [
            'NF', 'DATA_EXPEDICAO', 'DATA_CSV', 'STATUS', 'OPERAÇÃO', 
            'VOG_Expedição', 'PESO_CSV', 'R$ NF_Expedição', 'TOTAL_CSV'
        ]
        colunas_ordenadas = [col for col in colunas_ordenadas if col in df_relatorio.columns]
        df_relatorio = df_relatorio[colunas_ordenadas]
        
        downloads_path = str(Path.home() / "Downloads")
        caminho_relatorio = os.path.join(downloads_path, "RELATORIO_DIVERGENCIAS.xlsx")
        
        with pd.ExcelWriter(caminho_relatorio, engine='openpyxl') as writer:
            df_relatorio.to_excel(writer, index=False, sheet_name='Divergências')
            worksheet = writer.sheets['Divergências']
            
            date_columns = ['DATA_EXPEDICAO', 'DATA_CSV']
            for col_idx, col_name in enumerate(df_relatorio.columns):
                if col_name in date_columns:
                    col_letter = chr(65 + col_idx)
                    for row in range(2, len(df_relatorio) + 2):
                        cell = worksheet[f'{col_letter}{row}']
                        if cell.value and not pd.isna(cell.value):
                            cell.number_format = 'DD/MM/YYYY'
            
            for col_idx, col_name in enumerate(df_relatorio.columns):
                if col_name == 'NF':
                    col_letter = chr(65 + col_idx)
                    for row in range(2, len(df_relatorio) + 2):
                        cell = worksheet[f'{col_letter}{row}']
                        cell.number_format = '0'
            
            float_columns = ['PESO_CSV', 'TOTAL_CSV']
            for col_idx, col_name in enumerate(df_relatorio.columns):
                if col_name in float_columns:
                    col_letter = chr(65 + col_idx)
                    for row in range(2, len(df_relatorio) + 2):
                        cell = worksheet[f'{col_letter}{row}']
                        cell.number_format = '#,##0.00'
            
            aplicar_estilo_erros(worksheet, df_relatorio, problemas_por_linha)
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # RESUMO FINAL SIMPLIFICADO
        print(f"\n✅ RELATÓRIO CONCLUÍDO")
        print(f"📁 Salvo em: {caminho_relatorio}")
        print(f"📊 Total de divergências: {len(resultados)}")
        
        tipos_problemas = {}
        for problemas in problemas_por_linha.values():
            for problema in problemas:
                tipo = problema['tipo']
                tipos_problemas[tipo] = tipos_problemas.get(tipo, 0) + 1
        
        print("\n🔍 RESUMO DE PROBLEMAS:")
        if 'nf_nao_encontrada' in tipos_problemas:
            print(f"   🔴 NFs não encontradas: {tipos_problemas['nf_nao_encontrada']}")
        if 'peso_divergente' in tipos_problemas:
            print(f"   🔴 Divergências de PESO: {tipos_problemas['peso_divergente']}")
        if 'valor_divergente' in tipos_problemas:
            print(f"   🔴 Divergências de VALOR: {tipos_problemas['valor_divergente']}")
        if 'erro_digitacao_vog' in tipos_problemas:
            print(f"   🟡 Erros digitação VOG: {tipos_problemas['erro_digitacao_vog']}")
        if 'erro_digitacao_valor' in tipos_problemas:
            print(f"   🟡 Erros digitação VALOR: {tipos_problemas['erro_digitacao_valor']}")
            
    else:
        print("\n✅ Nenhuma divergência encontrada!")
    
    # Estatísticas rápidas
    nfs_sem_match_expedicao = len(df_comparacao[df_comparacao['_merge'] == 'left_only'])
    nfs_sem_match_csv = len(nfs_csv_sem_expedicao)
    
    print(f"\n📈 ESTATÍSTICAS:")
    print(f"   📋 Notas expedição VOG: {len(df_expedicao_filtrado)}")
    print(f"   📋 Notas CSV (histórico 51 + QTDE REAL positiva): {len(df_agrupado)}")
    print(f"   ❌ Expedição sem CSV (histórico 51 + QTDE REAL positiva): {nfs_sem_match_expedicao}")
    print(f"   ❌ CSV (histórico 51 + QTDE REAL positiva) sem expedição: {nfs_sem_match_csv}")

if __name__ == "__main__":
    processar_planilhas()
    input("\nPressione Enter para sair...")